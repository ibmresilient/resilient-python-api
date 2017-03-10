#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""Report users and groups into a spreadsheet"""

from __future__ import print_function
import os
import logging
import co3 as resilient
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font


# The config file location should usually be set in the environment
APP_CONFIG_FILE = os.environ.get("APP_CONFIG_FILE", "report.config")


def get_datetime(ts):
    """datetime from epoch timestamp"""
    if ts:
        return datetime.fromtimestamp(int(int(ts)/1000))


def report_users_and_groups(client, filename="users_groups.xlsx"):
    """Produce a spreadsheet report of the users and groups in this organization"""
    # Get the lists of users and groups
    users = client.get("/users?want_disabled=true")
    groups = client.get("/groups")

    # Make lookups for convenience
    users_lookup = {user["id"]: user["email"] for user in users}
    groups_lookup = {group["id"]: group["name"] for group in groups}

    # Sort users by last name
    users = sorted(users, key=lambda user: user["lname"] + user["fname"])
    # Sort groups by name
    groups = sorted(groups, key=lambda group: group["name"])

    # Create a spreadsheet that lists them all

    workbook = Workbook()
    ws_users = workbook.active
    ws_users.title = "Users"
    ws_groups = workbook.create_sheet(title="Groups")

    users_column_map = {
        "id": 1,
        "email": 2,
        "fname": 3,
        "lname": 4,
        "last_login": 5,
        "title": 6,
        "phone": 7,
        "enabled": 8,
        "status": 9,
        "group_ids": 14
    }
    users_role_column_map = {
        "administrator": 10,
        "create_incs": 11,
        "master_administrator": 12,
        "observer": 13
    }
    groups_column_map = {
        "id": 1,
        "name": 2,
        "members": 3
    }

    # Users sheet: make the header row
    row = 1
    for col in users_column_map:
        ws_users.cell(row=row, column=users_column_map[col], value=col).font = Font(bold=True)
    for col in users_role_column_map:
        ws_users.cell(row=row, column=users_role_column_map[col], value=col).font = Font(bold=True)

    # Write the user rows
    for user in users:
        row += 1
        for col in users_column_map:
            value = user.get(col)
            if col == "group_ids":
                # resolve names of all the groups
                user_groups = sorted([groups_lookup.get(group, "?") for group in value])
                value = ", ".join(user_groups)
            elif col == "last_login":
                # This is a time/date
                value = get_datetime(value)
            ws_users.cell(row=row, column=users_column_map[col], value=value)
        for col in users_role_column_map:
            value = user["roles"].get(col)
            value = "Y" if value else ""
            ws_users.cell(row=row, column=users_role_column_map[col], value=value)

    # Groups sheet: make the header row
    row = 1
    for col in groups_column_map:
        ws_groups.cell(row=row, column=groups_column_map[col], value=col).font = Font(bold=True)

    # Write the group rows
    for group in groups:
        row += 1
        for col in groups_column_map:
            value = group.get(col)
            if col == "members":
                # resolve names of all the users
                group_users = sorted([users_lookup.get(user, "?") for user in value])
                value = ", ".join(group_users)
            ws_groups.cell(row=row, column=groups_column_map[col], value=value)

    workbook.save(filename=filename)
    print("Written to {}".format(filename))


def main():
    """Main"""

    # Parse the commandline arguments and config file
    parser = resilient.ArgumentParser(config_file=APP_CONFIG_FILE)
    opts = parser.parse_args()

    # Create SimpleClient for a REST connection to the Resilient services
    resilient_client = resilient.get_client(opts)

    # Report the list of users and groups
    report_users_and_groups(resilient_client)


if __name__ == "__main__":
    logging.basicConfig(level=logging.WARN)
    main()
